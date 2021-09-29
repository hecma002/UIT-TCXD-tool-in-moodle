try:
        from openpyxl import Workbook, load_workbook
except:
        print ("openpyxl is missiong!")
        exit()

import sys
import os
import re
import codecs
import base64

#Checking file
if len(sys.argv) < 2:
        print ("Usage: python3 " + sys.argv[0], "source_file.xlsx"   )        
        sys.exit(1)

sourceFile = sys.argv[1]
#Check if source file is exist
if not os.path.exists(sourceFile):
        print ("Source file does not exist.")
        sys.exit(1)

wb = load_workbook(sourceFile)
ws = wb['auto-shuffle']

#Detect number of question

imgDir = "images/"

targetFile = re.sub("\.xlsx$", ".txt", sourceFile)
targetStream = codecs.open(targetFile, "w", 'utf8')
line = 3

attemptsIgnored = 0

def loadImage(name):
        name = name.strip()
        imgStr = ""
        try:
                with open(imgDir + name, "rb") as image_file:
                        imgStr = base64.b64encode(image_file.read()).decode('utf-8')
                        imgStr = re.sub("([=:])", r"\\\g<1>", imgStr)
                        if ".jpg" in name:
                                imgStr = '<img src\="data\:image/jpeg;base64,' + imgStr + '" />'
                        elif ".png" in name:
                                imgStr = '<img src\="data\:image/png;base64,' + imgStr + '" />'
        except:
                print("\tErr: image not found:", name)
        return imgStr

while True:
        questId = "0"*4 + str(line-2)
        questId = questId[-3:]

        questSummary = ws.cell(row=line, column = 2).value

        if(questSummary == None):
                attemptsIgnored += 1
                if(attemptsIgnored >= 3):
                        print ("Done")
                        break
                continue
        else:
                attemptsIgnored = 0

        print ("Processing", questId, "...")
        questSummary = re.sub("[:]+$", "", questSummary)
        questContent = ws.cell(row=line, column = 2).value

        #Find images in content
        imgs = re.findall("\[img:([^\]]+)\]", questContent)
        for img in imgs:
                questContent = questContent.replace("[img:"+img+"]", loadImage(img))

        correctAnswer = ws.cell(row=line, column = 3).value

        correctAnswer = str(correctAnswer)
        imgs = re.findall("\[img:([^\]]+)\]", correctAnswer)
        for img in imgs:
                correctAnswer = correctAnswer.replace("[img:"+img+"]", loadImage(img))
        

        answerString = "= " + correctAnswer + "\n"
        for i in range(4, 10):
                answer = ws.cell(row=line, column = i).value
                if answer:
                        answer = str(answer)
                        imgs = re.findall("\[img:([^\]]+)\]", answer)
                        for img in imgs:
                                answer = answer.replace("[img:"+img+"]", loadImage(img))
                        answerString += "~ " + answer + "\n"
        
        targetStream.write("::" + questId + ": " + questSummary + "::" + questContent + "\n")
        targetStream.write("{\n")
        targetStream.write(answerString)
        targetStream.write("}\n\n")

        line += 1


targetStream.close()
        